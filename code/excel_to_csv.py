#!/usr/bin/env python
# coding: utf-8

import pandas as pd
import os
from openpyxl import load_workbook

# make directory to put the csv files generated
if not os.path.exists('../data/survey_csv'):
    os.mkdir('../data/survey_csv')

# need the list of the raw excel files 
files = os.listdir('../data/survey_originals/')


# this function processes a file and extract the necessary sheet from an excel file
def process_file(file):
    '''
    Takes in an excel file from NYC Schools Surveys and generates the nessary information
    separated by student-, parent-, teacher-, or total(aggregated)-data and writes corresponding
    csv to file.
    '''
    # store survey year
    year = file.split('_')[0]
    
    # get studetn survey results
    if 'student' in file:
        # get the sheet name for the student survey results
        # inspiration from:
        # https://stackoverflow.com/questions/17977540/pandas-looking-up-the-list-of-sheets-in-an-excel-file
        sheets = load_workbook(f'../data/survey_originals/{file}', read_only=True).sheetnames
        for sheet in sheets:
            if 'Student %' in sheet:
                sheet_name = sheet

        student = pd.read_excel(f'../data/survey_originals/{file}', sheet_name=sheet_name, skiprows=[2], header=[0, 1])
        # write to file
        student.to_csv(f'../data/survey_csv/{year}_student.csv', index=False) 

    elif 'parent' in file:
        # get the sheet name for where the parent survey results are
        sheets = load_workbook(f'../data/survey_originals/{file}', read_only=True).sheetnames
        for sheet in sheets:
            if 'Parent %' in sheet:
                sheet_name = sheet

        parent = pd.read_excel(f'../data/survey_originals/{file}', sheet_name=sheet_name, skiprows=[2], header=[0, 1])
        # write to file
        parent.to_csv(f'../data/survey_csv/{year}_parent.csv', index=False)

    elif 'teacher' in file:
        sheets = load_workbook(f'../data/survey_originals/{file}', read_only=True).sheetnames
        for sheet in sheets:
            if 'Teacher %' in sheet:
                sheet_name = sheet
            elif 'Total' in sheet:
                tot_sheet_name = sheet

        teacher = pd.read_excel(f'../data/survey_originals/{file}', sheet_name=sheet_name, skiprows=[2], header=[0, 1])
        total = pd.read_excel(f'../data/survey_originals/{file}', sheet_name=tot_sheet_name)
        
        # write to file
        teacher.to_csv(f'../data/survey_csv/{year}_teacher.csv', index=False)
        total.to_csv(f'../data/survey_csv/{year}_total.csv', index=False)

    else:
        sheets = load_workbook(f'../data/survey_originals/{file}', read_only=True).sheetnames
        for sheet in sheets:
            if 'Student %' in sheet:
                st_sheet_name = sheet
            elif 'Parent %' in sheet:
                pa_sheet_name = sheet
            elif 'Teacher %' in sheet:
                te_sheet_name = sheet
            elif 'Total' in sheet:
                tot_sheet_name = sheet

        student = pd.read_excel(f'../data/survey_originals/{file}', sheet_name=st_sheet_name, skiprows=[2], header=[0, 1])
        parent = pd.read_excel(f'../data/survey_originals/{file}', sheet_name=pa_sheet_name, skiprows=[2], header=[0, 1])
        teacher = pd.read_excel(f'../data/survey_originals/{file}', sheet_name=te_sheet_name, skiprows=[2], header=[0, 1])
        total = pd.read_excel(f'../data/survey_originals/{file}', sheet_name=tot_sheet_name)
        
        # write to file
        student.to_csv(f'../data/survey_csv/{year}_student.csv', index=False) 
        teacher.to_csv(f'../data/survey_csv/{year}_teacher.csv', index=False)
        parent.to_csv(f'../data/survey_csv/{year}_parent.csv', index=False)
        total.to_csv(f'../data/survey_csv/{year}_total.csv', index=False)
        
        return 'Done'

for file in files:
    if file.endswith('xlsx'):
        process_file(file)